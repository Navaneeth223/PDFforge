import os
import openpyxl
import csv
import json
from openpyxl import Workbook
from typing import List
from services.storage import get_output_path

def excel_to_csv(session_id: str, file_path: str) -> str:
    output_path = get_output_path(session_id, ".csv")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sh = wb.active
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        col = csv.writer(f)
        for r in sh.rows:
            col.writerow([cell.value for cell in r])
    return output_path

def excel_to_json(session_id: str, file_path: str) -> str:
    output_path = get_output_path(session_id, ".json")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sh = wb.active
    data = []
    rows = list(sh.rows)
    header = [cell.value for cell in rows[0]]
    for row in rows[1:]:
        data.append(dict(zip(header, [cell.value for cell in row])))
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)
    return output_path

def merge_excel_sheets(session_id: str, file_paths: List[str]) -> str:
    output_path = get_output_path(session_id, ".xlsx")
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for path in file_paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            new_sheet = new_wb.create_sheet(title=f"{os.path.basename(path)}_{sheet.title}"[:30])
            for row in sheet.iter_rows():
                new_sheet.append([cell.value for cell in row])
    new_wb.save(output_path)
    return output_path
